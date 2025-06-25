import sqlite3
from datetime import datetime
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import tkinter as tk

def export_to_excel(treeview, columns):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты теста"

        for col_num, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=col_name).font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col_num)].width = len(col_name) + 2

        for row_num, item in enumerate(treeview.get_children(), 2):
            values = treeview.item(item)['values']
            for col_num, value in enumerate(values, 1):
                ws.cell(row=row_num, column=col_num, value=value)
                if columns[col_num - 1] in ("Сред. время реакции", "СКО", "Коэф. Уиппла"):
                    ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='right')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        filename = f"simple_reaction_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        messagebox.showinfo("Экспорт завершен", f"Данные успешно экспортированы в файл {filename}")
    except Exception as e:
        messagebox.showerror("Ошибка экспорта", f"Не удалось экспортировать данные: {str(e)}")

def show_help_window(parent):
    help_window = tk.Toplevel(parent)
    help_window.title("Справка по таблице результатов")
    help_window.geometry("800x600")

    help_text = """
      Функции таблицы результатов:

      1. Сортировка данных:
         - Нажмите на заголовок столбца для сортировки по возрастанию
         - Нажмите еще раз для сортировки по убыванию
         - ФИО и интерпретация сортируются в алфавитном порядке при нажатии на название столбца

      2. Поиск данных:
         - По ФИО: можно вводить часть имени, фамилии или отчества
         - По дате: можно вводить часть даты в формате ДД.ММ.ГГГГ
         - Можно вводить несколько значений через запятую

      3. Экспорт данных:
         - Нажмите кнопку "Экспорт в Excel" для сохранения таблицы в файл Excel

      4. Статистика:
         - В верхней части окна отображается средняя статистика 
         - Можно отобразить лучшие результаты или результаты выше среднего

      5. Редактирование:
         - Двойной клик по ячейке "Интерпретация" позволяет редактировать данные
         - Нажмите Enter для сохранения или Esc для отмены

    """

    text_widget = tk.Text(help_window, wrap=tk.WORD, font=('Helvetica', 12))
    text_widget.insert(tk.END, help_text)
    text_widget.config(state=tk.DISABLED)
    text_widget.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

    close_button = tk.Button(help_window, text="Закрыть", command=help_window.destroy, font=('Helvetica', 13, 'bold'))
    close_button.pack(pady=10)

